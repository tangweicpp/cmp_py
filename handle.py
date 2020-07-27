'''
@File    :   handle.py
@Time    :   2020/07/09 15:30:00
@Author  :   Tang wei
@Version :   1.0
@Contact :   wei.tang_ks@ht-tech.com
@License :   (C)Copyright 2020-2021
@Desc    :   None
'''
import connect_db as conn
import time
import os
import send_email as se
import pandas as pd
import numpy as np
import json
import re
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl import load_workbook


upload_task = {}
os.environ['NLS_LANG'] = 'SIMPLIFIED CHINESE_CHINA.UTF8'


# None to ''
def xstr(s):
    return '' if s is None else str(s).strip()


# Upload progress
def get_progress(user_key):
    global upload_task

    if user_key in upload_task:
        return upload_task[user_key]
    return 0


# Check username and password
def check_account(username, password):
    if not (username and password):
        print('用户名或密码为空')
        return False

    sql = "select 用户号 from erpbase..tblOperatorData where 用户号 = '%s' and 密码= '%s' " % (
        username, password)
    results = conn.MssConn.query(sql)
    if not results:
        print('查询不到数据')
        return False
    return True


# Get customer list
def get_cust_code_list():
    json_data = []

    sql = "SELECT DISTINCT CUSTOMERSHORTNAME FROM TBLTSVNPIPRODUCT ORDER BY CUSTOMERSHORTNAME "
    results = conn.OracleConn.query(sql)
    for row in results:
        result = {}
        result['value'] = xstr(row[0])
        result['label'] = xstr(row[0])

        json_data.append(result)
    return json_data


# Get customer po template
def get_po_template(cust_code):
    if not cust_code:
        print('客户代码不可为空')
        return []

    json_data = []
    sql = "SELECT CUST_CODE,TEMPLATE_FILE ,TEMPLATE_PIC ,KEY_LIST ,FILE_LEVEL,FILE_URL,ACCEPT,TEMPLATE_ID FROM CMP_CUST_PO_TEMPLATE WHERE CUST_CODE  = '%s'" % (
        cust_code)
    results = conn.OracleConn.query(sql)
    for row in results:
        result = {}
        result['file_name'] = xstr(row[1])
        result['img_url'] = xstr(row[2])
        result['file_key'] = xstr(row[3])
        result['level'] = xstr(row[4])
        result['file_url'] = xstr(row[5])
        result['accept'] = xstr(row[6])
        result['file_id'] = xstr(row[7])
        result['show_progress_flag'] = False
        result['show_filelist_flag'] = False
        result['load_progress'] = 0

        json_data.append(result)
    return json_data


# Upload po file
def upload_po_file(f, po_header):
    global upload_task
    upload_task[po_header['file_id']] = 0

    if not f:
        print('文件不存在')
        po_header['err_desc'] = '上传的文件不存在'
        return False

    file_dir = os.path.join(os.getcwd(), 'uploads/po/' +
                            po_header['po_type']+'/'+po_header['cust_code'])

    if not os.path.exists(file_dir):
        os.makedirs(file_dir)

    file_path = os.path.join(file_dir, f.filename)
    f.save(file_path)

    sql = "select PO_ITEM_SEQ.nextval from dual"
    po_header['upload_id'] = conn.OracleConn.query(sql)[0][0]

    if not parse_po_file(file_path, po_header):
        return False

    ret = get_upload_data(po_header['upload_id'])
    mail_attachment = []
    mail_attachment.append(file_path)
    mail_attachment.append(os.path.join(os.getcwd(), '已上传订单.xlsx'))
    send_mail(ret, po_header, mail_attachment)

    upload_task[po_header['file_id']] = 100

    return ret


# Sent mail
def send_mail(ret_data, po_header, mail_attachment):
    mail_keyid = po_header['upload_id']

    mail_body = get_mail_body(
        po_header['user_name'], mail_keyid, po_header['mail_tip'], ret_data)

    sql = "select recv_user_to from erp_email_recv where email_type = 'WO_UPLOAD_RECV_TEST' "
    mail_recv = conn.OracleConn.query(sql)[0][0].split(',')

    sql = "select recv_user_cc from erp_email_recv where email_type = 'WO_UPLOAD_RECV_TEST' "
    mail_recv_cc = conn.OracleConn.query(sql)[0][0].split(',')

    mail_title = '测试新版WO上传'

    se.send_email(mail_title, mail_body, mail_attachment,
                  mail_recv, mail_recv_cc)


# Get mail body
def get_mail_body(createBy, uploadid, addContent, json_data):
    # Get total data
    content = f'''
          <h2>内勤人员：{createBy} 上传WO</h2>
          <div style="color:red"><pre>{addContent}</pre></div>
          <h2>以下为上传WO的汇总信息</h2>
          <table border="1" cellspacing="0" cellpadding="5" style="border:rgb(175, 175, 175) solid thin">
              <tr style="background-color: rgb(175, 175, 175);">
                  <th>序号</th><th>保税</th><th>客户代码</th><th>客户PO</th><th>客户机种</th><th>客户Fab机种</th><th>厂内机种</th><th>晶圆料号</th>
                  <th>客户LOTID</th><th>片数</th><th>DIES</th><th>上传人员工号</th><th>上传时间</th>
              </tr>
          '''

    for row in json_data['total_data']:
        content = content + f'''<tr><td>{row['id']}</td><td>{row['banded']}</td><td>{row['cust_code']}</td><td>{row['po_id']}</td>
        <td>{row['cust_device']}</td><td>{row['fab_device']}</td><td>{row['ht_pn']}</td><td>{row['wafer_pn']}</td>
        <td>{row['lot_id']}</td><td>{row['wafer_qty']}</td><td>{row['die_qty']}</td><td>{row['upload_by']}</td><td>{row['upload_date']}</td></tr>
        '''

    content = content + '</table>'

    # Get detail data
    content = content + f'''<h2>以下为明细数据</h2>
    <table border="1" cellspacing="0" cellpadding="5" style="border:rgb(175, 175, 175) solid thin">
              <tr style="background-color: rgb(175, 175, 175);">
                  <th>序号</th><th>客户PO</th><th>客户机种</th><th>客户Fab机种</th><th>厂内机种</th>
                  <th>客户LOTID</th><th>waferID</th><th>grossdies</th><th>打标码</th>
              </tr>
    '''

    for row in json_data['detail_data']:
        content = content + f'''<tr><td>{row['id']}</td><td>{row['po_id']}</td>
        <td>{row['cust_device']}</td><td>{row['fab_device']}</td><td>{row['ht_pn']}</td>
        <td>{row['lot_id']}</td><td>{row['wafer_id']}</td><td>{row['gross_dies']}</td><td>{row['mark_code']}</td></tr>
        '''

    content = content + '</table>'

    return content


# Get upload data
def get_upload_data(upload_id):
    dict_data = {}
    json_data_total = []
    json_data_detail = []

    # Total data
    sql = f'''
    select row_number() over(ORDER BY bb.lotid) as 序号,case bb.substratetype when 'A' THEN '保税' when 'B' THEN '非保税' else '未知' end as 是否保税,
    bb.customershortname as 客户代码,aa.po_num,aa.mpn_desc as 客户机种名,aa.Fab_conv_id as 客户Fab机种,aa.mtrl_num as 厂内机种,cc.MARKETLASTUPDATE_BY AS 晶圆料号,
    bb.lotid AS LOTID,count(bb.wafer_id) AS 片,sum(bb.passbincount + bb.failbincount) Dies,bb.qtech_created_by, to_char(bb.qtech_created_date,'yyyy-MM-dd'),cc.residual
    from customeroitbl_test aa inner join mappingdatatest bb on to_char(aa.id) = bb.filename INNER JOIN TBLTSVNPIPRODUCT cc ON aa.CUSTOMERSHORTNAME  = cc.CUSTOMERSHORTNAME
    AND aa.MPN_DESC  = cc.CUSTOMERPTNO1 AND aa.FAB_CONV_ID  = cc.CUSTOMERPTNO2  where aa.wafer_visual_inspect = '{upload_id}'
    group by bb.customershortname,aa.Fab_conv_id,aa.mpn_desc,bb.lotid,aa.mtrl_num,bb.passbincount,bb.failbincount,aa.po_num,bb.qtech_created_by,
    to_char(bb.qtech_created_date,'yyyy-MM-dd'),bb.substratetype,cc.MARKETLASTUPDATE_BY,cc.residual
    '''
    results = conn.OracleConn.query(sql)
    for row in results:
        result = {}
        result['id'] = xstr(row[0])
        result['banded'] = xstr(row[1])
        result['cust_code'] = xstr(row[2])
        result['po_id'] = xstr(row[3])
        result['cust_device'] = xstr(row[4])
        result['fab_device'] = xstr(row[5])
        result['ht_pn'] = xstr(row[6])
        result['wafer_pn'] = xstr(row[7])
        result['lot_id'] = xstr(row[8])
        result['wafer_qty'] = row[9]
        result['die_qty'] = row[10]
        result['upload_by'] = xstr(row[11])
        result['upload_date'] = xstr(row[12])
        result['npi_owner'] = xstr(row[13])

        json_data_total.append(result)

    dict_data['total_data'] = json_data_total

    # Detail data
    sql = f'''
    select row_number() over(ORDER BY  bb.lotid,bb.substrateid) as 序号,case bb.substratetype when 'A' then '保税' when 'B' then '非保税' else '未知' end as 是否保税, bb.customershortname as 客户代码,
          aa.Fab_conv_id as FAB机种,aa.mpn_desc as 客户机种,cc.residual as NPI负责人员,
          aa.mtrl_num as 厂内机种,
          aa.po_num as PO_NUM,
          bb.lotid as LOT_ID,
          bb.wafer_id as WAFER_NO,
          bb.substrateid as WAFERID,
          bb.passbincount as GOOD_DIES,
          bb.failbincount as NG_DIES,
          bb.passbincount + bb.failbincount as GROSS_DIES,
          bb.productid as 打标码,
          aa.Imager_Customer_Rev as 二级代码, bb.qtech_created_by as 上传人员,bb.qtech_created_date as 上传时间,  bb.qtech_lastupdate_by as 更新人员, bb.qtech_lastupdate_date as 更新时间
     from customeroitbl_test aa
     left join tbltsvnpiproduct cc on cc.customerptno1 = aa.mpn_desc  and  cc.qtechptno = aa.mtrl_num  and cc.customershortname = aa.customershortname and cc.residual is not null
    inner join mappingdatatest bb
       on to_char(aa.id) = bb.filename
      and aa.wafer_visual_inspect = '{upload_id}'
      group by  bb.customershortname,cc.residual,aa.mtrl_num,aa.Fab_conv_id, aa.mpn_desc,aa.po_num,bb.lotid,bb.wafer_id,bb.substrateid,bb.passbincount,bb.failbincount,bb.productid,aa.Imager_Customer_Rev ,bb.substratetype,bb.qtech_created_by,bb.qtech_created_date,bb.qtech_lastupdate_by,bb.qtech_lastupdate_date
    '''

    results = conn.OracleConn.query(sql)
    for row in results:
        result = {}
        result['id'] = xstr(row[0])
        result['banded'] = xstr(row[1])
        result['cust_code'] = xstr(row[2])
        result['fab_device'] = xstr(row[3])
        result['cust_device'] = xstr(row[4])
        result['npi_owner'] = xstr(row[5])
        result['ht_pn'] = xstr(row[6])
        result['po_id'] = xstr(row[7])
        result['lot_id'] = xstr(row[8])
        result['wafer_no'] = xstr(row[9])
        result['wafer_id'] = xstr(row[10])
        result['good_dies'] = row[11]
        result['ng_dies'] = row[12]
        result['gross_dies'] = row[13]
        result['mark_code'] = xstr(row[14])
        result['second_code'] = xstr(row[15])
        result['upload_by'] = xstr(row[16])
        result['upload_date'] = xstr(row[17])
        result['update_by'] = xstr(row[18])
        result['update_date'] = xstr(row[19])

        json_data_detail.append(result)

    dict_data['detail_data'] = json_data_detail

    set_xl(json_data_total, json_data_detail)
    return dict_data


# Get cell value
def get_cell_val(row, col, data, header):
    if col > len(header) or (row-2) > len(data):
        val = ""
    else:
        val = data[row-3][header[col-1]]
    return val


# Set new xl
def set_xl(total_data, detail_data):
    wb = load_workbook('export_xl_template/template.xlsx')
    # Total
    header_list = ["id", "po_id", "cust_code",  "cust_device", "fab_device",
                   "ht_pn", "lot_id", "wafer_qty", "die_qty", "banded", "upload_by", "upload_date", "npi_owner"]

    ws = wb.get_sheet_by_name(wb.sheetnames[0])
    for row in range(3, len(total_data)+3):
        for col in range(1, len(header_list)):
            ws.cell(column=col, row=row, value=get_cell_val(
                row, col, total_data, header_list))

    # Detail
    header_list = ["id", "po_id", "cust_code",  "cust_device", "fab_device", "ht_pn", "lot_id",
                   "wafer_no", "wafer_id", "gross_dies", "good_dies", "ng_dies", "mark_code"]

    ws = wb.get_sheet_by_name(wb.sheetnames[1])

    for row in range(3, len(detail_data)+3):
        for col in range(1, 30):
            ws.cell(column=col, row=row, value=get_cell_val(
                row, col, detail_data, header_list))

    wb.save("已上传订单.xlsx")


# Parse po file
def parse_po_file(file_name, po_header):
    po_dict = get_po_config(po_header)
    if not po_dict:
        po_header['err_desc'] = '查询不到配置文件无法解析'
        return False

    file_type = po_dict['file_type']
    if not file_name.endswith(file_type):
        po_header['err_desc'] = '您上传的文件和模板设定的文件类型不一致'
        return False

    if file_type == 'xlsx':
        if not parse_xlsx_file(file_name, po_header, po_dict):
            return False
    else:
        po_header['err_desc'] = '文件类型不支持解析'
        return False

    return True


# Get Json config
def get_po_config(po_header):
    sql = "SELECT TEMPLATE_CONFIG FROM CMP_CUST_PO_TEMPLATE WHERE TEMPLATE_ID  = %s" % (
        po_header['file_id'])
    results = conn.OracleConn.query(sql)
    if not results:
        print("无法获取配置文件")
        return False

    template_config = results[0][0]
    file_dir = os.path.join(os.getcwd(), template_config)

    f = open(file_dir, 'r', encoding="utf-8")
    po_dict = json.load(f)
    return po_dict


# Parse xlsx file
def parse_xlsx_file(file_name, po_header, po_dict):
    file_cfg = ['file_key', 'file_index', 'file_header', 'file_max_cols']
    for item in file_cfg:
        if not item in po_dict:
            po_header['err_desc'] = '配置文件错误，请联系IT处理'
            return False

    file_key = po_dict['file_key']
    file_index = po_dict['file_index']
    file_header = po_dict['file_header']
    file_max_cols = po_dict['file_max_cols']

    po_data = []
    df = pd.DataFrame(pd.read_excel(
        file_name, sheet_name=file_index, header=file_header, keep_default_na=False))
    for index, row in df.iterrows():
        po_row_data = {}
        for key in file_key:
            col_name = file_key[key]['position']['col_name']
            po_row_data[key] = row[col_name] if col_name else ''

        po_data.append(po_row_data)

    print(po_data)
    if not check_po_data(po_header, po_dict, po_data):
        return False

    if not save_po_data(po_header, po_dict, po_data):
        return False
    return True


# Get list
def get_wafer_list(wafer_str):
    if wafer_str == "":
        return []

    wafer_str_new = re.sub(r'[_~-]', ' _ ', wafer_str)
    pattern = re.compile(r'[A-Za-z0-9_]+')
    result1 = pattern.findall(wafer_str_new)

    # extend
    for i in range(1, len(result1)-1):
        if result1[i] == '_':
            if result1[i-1].isdigit() and result1[i+1].isdigit():
                bt = []
                if int(result1[i-1]) < int(result1[i+1]):
                    for j in range(int(result1[i-1])+1, int(result1[i+1])):
                        bt.append(f'{j}')
                else:
                    for j in range(int(result1[i-1])-1, int(result1[i+1]), -1):
                        bt.append(f'{j}')
                result1.extend(bt)

    # remove '_'
    result2 = sorted(set(result1), key=result1.index)
    if '_' in result2:
        result2.remove('_')

    return result2


# Check po data
def check_po_data(po_header, po_dict, po_data):

    for item in po_data:
        # Check 0:necessary key
        necessary_key_list = ['po_id', 'customer_device', 'lot_id', 'wafer_id']
        for key in necessary_key_list:
            if not key in item:
                po_header['err_desc'] = '配置文件必要KEY不存在错误，无法解析上传文件，请联系IT处理'
                return False

        wafer_id_list = get_wafer_list(item['wafer_id'])
        if len(wafer_id_list) == 0:
            continue

        # Check 1:Wafer qty
        if 'wafer_qty' in item:
            wafer_qty = item['wafer_qty']
            if len(wafer_id_list) != int(wafer_qty):
                print('wafer qty和wafer list明细不一致')
                po_header['err_desc'] = 'wafer qty和wafer list明细不一致'
                return False

        # Check 2:

    return True


# Save po data
def save_po_data(po_header, po_dict, po_data):
    global upload_task

    num = 0
    for item in po_data:
        wafer_id_list = get_wafer_list(item['wafer_id'])
        num = num + len(wafer_id_list)

    for item in po_data:
        wafer_id_list = get_wafer_list(item['wafer_id'])
        if len(wafer_id_list) == 0:
            continue

        for i in range(len(wafer_id_list)):
            insert_po_data(wafer_id_list[i], po_header, item)
            upload_task[po_header['file_id']
                        ] = upload_task[po_header['file_id']] + 100 / float(num)
            if upload_task[po_header['file_id']] >= 100:
                upload_task[po_header['file_id']] = 99

    return True


# Insert to DB
def insert_po_data(wafer_id, po_header, po_data):
    # Get data
    if wafer_id.isdigit() and (len(wafer_id) == 1):
        wafer_id = ('00' + wafer_id)[-2:]
    sql = "select CustomerBCtbl_SEQ.nextval ID from dual"
    max_id = conn.OracleConn.query(sql)[0][0]
    upload_id = po_header['upload_id']
    bonded = 'A' if po_header['bonded_type'] == '保税' else 'B'
    create_by = po_header['user_name']
    lot_id = po_data['lot_id']
    cust_code = po_header['cust_code']
    po_id = po_data['po_id']
    cust_device = po_data['customer_device']
    fab_device = po_data['fab_device']
    ret = get_cust_pn_info(cust_device, fab_device)
    ht_pn = ret['ht_pn'] if ret else ''
    passbin_count = ret['gross_dies'] if ret else ''
    failbin_count = '0'
    product_id = ret['product_id'] if ret else ''

    ship_comment = po_data['add_1']
    probe_ship_part_type = po_data['add_2']
    reticle_level_71 = po_data['add_3']
    reticle_level_72 = po_data['add_4']
    reticle_level_73 = po_data['add_5']
    assembly_facility = po_data['add_6']
    batch_comment_assy = po_data['add_7']
    mark_code = po_data['mark_code']

    # Delete old wafer data
    delete_po_data('1', lot_id+wafer_id)
    # Oracle insert
    sql = ''' insert into mappingDataTest(id,substrateid,substratetype,productid,micronlotid,lotid,Wafer_ID,passbincount,
              failbincount,CustomerShortName,flag,Qtech_Created_By,Qtech_Created_Date,filename)
              values( mappingData_SEQ.Nextval,'%s','%s','%s','%s',
                     '%s','%s','%s','%s','%s','Y','%s',sysdate,'%s')
          ''' % (lot_id+wafer_id, bonded, mark_code, 'std_new', lot_id, wafer_id, passbin_count, failbin_count, cust_code, create_by, max_id)

    conn.OracleConn.exec(sql)

    sql = f''' insert into CustomerOItbl_test(id,po_num,wafer_visual_inspect,source_batch_id,SHIP_SITE,Test_site,FAB_CONV_ID,mpn_desc,
            Imager_Customer_Rev,Created_Date,mtrl_num,CustomerShortName,flag,Qtech_Created_By,Qtech_Created_Date,
            probe_ship_part_type,RETICLE_LEVEL_71,RETICLE_LEVEL_72,RETICLE_LEVEL_73,ASSEMBLY_FACILITY,BATCH_COMMENT_ASSY,
            jobno,date_code,shipping_mst_level,shipping_mst_260,TARGET_WAF_THICKNESS,COMP_CODE,SHIP_COMMENT)
            values( {max_id},'{po_id}','{upload_id}','{lot_id}','{cust_code}','HTKS','{fab_device}','{cust_device}','','','{ht_pn}','{cust_code}',
            'Y','{create_by}',sysdate,'{probe_ship_part_type}','{reticle_level_71}','{reticle_level_72}','{reticle_level_73}','{assembly_facility}',
            '{batch_comment_assy}','{bonded}','','','','{lot_id}','HTKS','{ship_comment}')
          '''

    conn.OracleConn.exec(sql)

    # Sqlserver insert
    sql = ''' insert into [ERPBASE].[dbo].[tblmappingData](substrateid,substratetype,productid,lotid,Wafer_ID,passbincount,
              failbincount,CustomerShortName,flag,Qtech_Created_By,Qtech_Created_Date,filename)
              values('%s','%s','%s','%s','%s','%s','0','%s','Y','%s',getdate(),'%s')
          ''' % (lot_id+wafer_id, bonded, mark_code, lot_id, wafer_id, passbin_count, cust_code, create_by, max_id)

    conn.MssConn.exec(sql)

    sql = f''' insert into [ERPBASE].[dbo].[tblCustomerOI](id,po_num,wafer_visual_inspect,source_batch_id,SHIP_SITE,Test_site,FAB_CONV_ID,mpn_desc,
            Imager_Customer_Rev,Created_Date,mtrl_num,CustomerShortName,flag,Qtech_Created_By,Qtech_Created_Date,
            probe_ship_part_type,RETICLE_LEVEL_71,RETICLE_LEVEL_72,RETICLE_LEVEL_73,ASSEMBLY_FACILITY,BATCH_COMMENT_ASSY,
            jobno,date_code,shipping_mst_level,shipping_mst_260,TARGET_WAF_THICKNESS,COMP_CODE,SHIP_COMMENT)
            values( {max_id},'{po_id}','{upload_id}','{lot_id}','{cust_code}','HTKS','{fab_device}','{cust_device}','','','{ht_pn}','{cust_code}',
            'Y','{create_by}',getdate(),'{probe_ship_part_type}','{reticle_level_71}','{reticle_level_72}','{reticle_level_73}','{assembly_facility}',
            '{batch_comment_assy}','{bonded}','','','','{lot_id}','HTKS','{ship_comment}')
          '''
    conn.MssConn.exec(sql)


def get_cust_pn_info(cust_device, fab_device):
    sql = "SELECT QTECHPTNO,CUSTOMERDIEQTY,QTECHPTNO2 FROM TBLTSVNPIPRODUCT t  WHERE t.CUSTOMERPTNO1  = '%s' AND CUSTOMERPTNO2  = '%s'" % (
        cust_device, fab_device)
    results = conn.OracleConn.query(sql)

    if not results:
        print('无法获取NPI对照表对应机种信息')
        return False

    ret = {}
    ret['ht_pn'] = results[0][0]
    ret['gross_dies'] = results[0][1]
    ret['product_id'] = results[0][2]
    return ret


# Delete po data
def delete_po_data(flag_, del_id):
    if flag_ == '0':
        # Delete by lot
        sql = f"delete from mappingdatatest where lotid = '{del_id}' "
        conn.OracleConn.exec(sql)
        sql = f"delete from CustomerOItbl_test where source_batch_id = '{del_id}' "
        conn.OracleConn.exec(sql)
        sql = f"delete from [ERPBASE].[dbo].[tblmappingData] where lotid = '{del_id} "
        conn.MssConn.exec(sql)
        sql = f"delete from [ERPBASE].[dbo].[tblCustomerOI] where source_batch_id = '{del_id}' "
        conn.MssConn.exec(sql)
    elif flag_ == '1':
        # Delete by wafer
        sql = f"delete from CustomerOItbl_test where to_char(id) in (select filename from mappingdatatest where substrateid = '{del_id}') "
        conn.OracleConn.exec(sql)
        sql = f"delete from mappingdatatest where substrateid = '{del_id}' "
        conn.OracleConn.exec(sql)
        sql = f"delete from [ERPBASE].[dbo].[tblCustomerOI] where CONVERT(char(20),id) in (select filename from [ERPBASE].[dbo].[tblmappingData] where substrateid = '{del_id}') "
        conn.MssConn.exec(sql)
        sql = f"delete from [ERPBASE].[dbo].[tblmappingData] where substrateid = '{del_id}' "
        conn.MssConn.exec(sql)
    elif flag_ == '2':
        # Delete by uploadid
        sql = f"delete from mappingdatatest where filename in (select to_char(id) from CustomerOItbl_test where wafer_visual_inspect = '{del_id}') "
        conn.OracleConn.exec(sql)
        sql = f"delete from CustomerOItbl_test where wafer_visual_inspect = '{del_id}' "
        conn.OracleConn.exec(sql)
        sql = f"delete from [ERPBASE].[dbo].[tblmappingData] where filename in (select CONVERT(char(20),id)  from [ERPBASE].[dbo].[tblCustomerOI] where wafer_visual_inspect = '{del_id}') "
        conn.MssConn.exec(sql)
        sql = f"delete from [ERPBASE].[dbo].[tblCustomerOI] where wafer_visual_inspect = '{del_id}' "
        conn.MssConn.exec(sql)
