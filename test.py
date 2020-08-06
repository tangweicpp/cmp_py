from handle import delete_po_data
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl import load_workbook
from itertools import groupby
import connect_db as conn

import re
import xlrd


# sql = '''
# select substrateid from mappingdatatest where micronlotid = 'std_new' order by substrateid 
# '''

# results = conn.OracleConn.query(sql)
# for row in results:
#     waferid = row[0]
#     conn.OracleConn.exec(
#         f"update mappingdatatest set QTECH_CREATED_DATE = sysdate where substrateid = '{waferid}'  ")

#     mark_id = conn.OracleConn.query(
#         "SELECT QTMCodeSeq.SXCode('a') FROM DUAL ")[0][0]

#     conn.OracleConn.exec(
#         f"update mappingdatatest set productid = '{mark_id}' where substrateid = '{waferid}'")
#     conn.MssConn.exec(
#         f"update ERPBASE.dbo.tblmappingData set productid = '{mark_id}' where substrateid = '{waferid}'")

# def thans_col_row_from_string(s):
#     dict = {}
#     s = 'AB10'
#     ss = [''.join(list(g)) for k, g in groupby(s, key=lambda x: x.isdigit())]

#     dict['col'] = column_index_from_string(ss[0]) - 1
#     dict['row'] = int(ss[1]) - 1
#     return dict


# str1 = ['K1']
# dict1 = thans_col_row_from_string(str1)
# print(dict1)


# ```
# 13105
# 13106
# 13107
# 13108
# 13109
# ```

# test_dict = {"abc": 123, "bcd": 234}
# for item in test_dict:
#     print(item)


sql = '''
SELECT DISTINCT t2.WAFER_VISUAL_INSPECT FROM MAPPINGDATATEST t1
INNER JOIN CUSTOMEROITBL_TEST t2 ON to_char(t2.id) = t1.FILENAME  
WHERE t1.MICRONLOTID = 'std_new' AND t1.LOTID LIKE 'TW%'
'''
results = conn.OracleConn.query(sql)
if results:
    del_list = results

    for del_id in del_list:
        delete_po_data('2', del_id[0])

        print(f'{del_id}删除成功')


# Get cell value by openpyxl

# key_postion = 'K1'
# file_name = 'GULF20034KS 7.22.xls'
# # wb = load_workbook(file_name)
# # ws = wb.get_sheet_by_name(wb.sheetnames[0])
# # cell_val = ws[key_postion].value
# # print(cell_val)


# data = xlrd.open_workbook(file_name)
# table = data.sheets()[0]
# column_index_from_string('K')

# data = table.cell_value(1-1, column_index_from_string('K')-1)
# # data = table.row_values(0)
# print(data)


# Get list


# def get_wafer_list(wafer_str):
#     if wafer_str == "":
#         return []

#     wafer_str_new = re.sub(r'[_~-]', ' _ ', wafer_str)
#     pattern = re.compile(r'[A-Za-z0-9_]+')
#     result1 = pattern.findall(wafer_str_new)

#     # extend
#     for i in range(1, len(result1)-1):
#         if result1[i] == '_':
#             if result1[i-1].isdigit() and result1[i+1].isdigit():
#                 bt = []
#                 if int(result1[i-1]) < int(result1[i+1]):
#                     for j in range(int(result1[i-1])+1, int(result1[i+1])):
#                         bt.append(f'{j}')
#                 else:
#                     for j in range(int(result1[i-1])-1, int(result1[i+1]), -1):
#                         bt.append(f'{j}')
#                 result1.extend(bt)

#     # remove '_'
#     result2 = sorted(set(result1), key=result1.index)
#     if '_' in result2:
#         result2.remove('_')

#     return result2


# get_wafer_list('01')
