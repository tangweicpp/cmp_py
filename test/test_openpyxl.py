from openpyxl.utils import get_column_letter, column_index_from_string
import openpyxl

# 1. 读取excel文档
wb = openpyxl.load_workbook('AP623003P_20200525_Bump_Prod.xlsx')
# print(wb, type(wb))


# 2. 在工作薄中取得工作表
# print(wb.get_sheet_names())
# 返回一个列表， 存储excel表中所有的sheet工作表;
# print(wb.sheetnames)

# 返回一个worksheet对象， 返回当前的活动表;
# print(wb.get_active_sheet())
# print(wb.active)


# 3. 获取工作表中， 单元格的信息
# sheet = wb[0]
# print(sheet['A1'])


sheetnames = wb.get_sheet_names()  # 获得表单名字
sheet = wb.get_sheet_by_name(sheetnames[0])  # 从工作表中提取某一表单



# sheet._charts.
# print(sheet['A1'].value)

# print(sheet['B1'].value)

# print(sheet['H2'].value)


# cell = sheet['H2']
# print(cell.row, cell.column, cell.coordinate)


# print(sheet.max_column)
# print(sheet.max_row)
# print(sheet.title)
# print(len(sheet._images))
# print(sheet._images[0].anchor._from.row, sheet._images[0].anchor._from.col)
# print(sheet['D8'].value)

# sheet.rows为生成器, 里面是每一行的数据，每一行又由一个tuple包裹。
# sheet.columns类似，不过里面是每个tuple是每一列的单元格。
# for row in sheet.rows:
#     for cell in row:
#         print(cell.value)


# for column in sheet.columns:
#     for cell in column:
#         print(cell.value)


# 根据列的数字返回字母
print(get_column_letter(27))  # B
# 根据字母返回列的数字
print(column_index_from_string('AA'))  # 4
